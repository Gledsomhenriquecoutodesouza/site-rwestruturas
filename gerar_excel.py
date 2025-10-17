import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image



def gerar_planilha_profissional():
    # --- CONFIGURAÇÕES BÁSICAS ---
    filename = "Proposta_Hospedagem_RW.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Opções de Hospedagem"

    # --- DEFINIÇÃO DE ESTILOS ---
    # Cores da marca
    cor_dourada = "D4AF37"
    cor_preta = "121212"
    cor_fundo_destaque = "FFF8DC"  # Um dourado bem claro

    # Estilos de Célula
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color=cor_preta, end_color=cor_preta, fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    title_font = Font(name='Calibri', size=16, bold=True)
    title_alignment = Alignment(horizontal='center', vertical='center')

    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    highlight_fill = PatternFill(start_color=cor_fundo_destaque, end_color=cor_fundo_destaque, fill_type='solid')
    bold_font = Font(bold=True)

    # --- INSERÇÃO DO LOGO E TÍTULO ---
    try:
        logo = Image('logo_rw.png')
        logo.height = 75
        logo.width = 160
        ws.add_image(logo, 'A1')
    except FileNotFoundError:
        print("AVISO: Arquivo 'logo_rw.png' não encontrado. O logo não será adicionado.")

    ws.merge_cells('C1:G2')
    title_cell = ws['C1']
    title_cell.value = "Comparativo de Opções de Hospedagem - RW Estruturas"
    title_cell.font = title_font
    title_cell.alignment = title_alignment

    # --- DADOS DA TABELA ---
    # Adicionamos \n para quebras de linha e "• " para bullet points
    data = [
        ['Opção', 'Endereço do Site (URL)', 'Custo do Domínio (.com.br)', 'Custo da Hospedagem',
         'Custo Total Anual (Estimado)', 'Vantagens', 'Desvantagens'],
        ['100% Gratuito', 'exemplo.pythonanywhere.com', 'R$ 0,00', 'R$ 0,00', 'R$ 0,00',
         '• Custo zero\n• Ótimo para testes internos',
         '• Endereço não profissional\n• Site "adormece" (lento para o primeiro visitante)\n• Menos potente'],
        ['Híbrido (Domínio Profissional)', 'www.rwestruturas.com.br', 'R$ 40,00 / ano', 'R$ 0,00', 'R$ 40,00',
         '• Endereço profissional\n• Custo anual muito baixo',
         '• O site também "adormece"\n• A performance pode ser limitada'],
        ['Profissional (Recomendado)', 'www.rwestruturas.com.br', 'R$ 40,00 / ano', '~ R$ 30,00 / mês', '~ R$ 400,00',
         '• Máxima credibilidade e profissionalismo\n• Alta performance (site sempre rápido)\n• Ideal para negócios e imagem da marca',
         '• Requer um pequeno investimento mensal']
    ]

    # --- PREENCHIMENTO E FORMATAÇÃO DA PLANILHA ---
    start_row = 4  # Linha onde a tabela começa
    for i, row_data in enumerate(data):
        for j, cell_data in enumerate(row_data):
            cell = ws.cell(row=start_row + i, column=j + 1, value=cell_data)

            # Formatação do cabeçalho
            if i == 0:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            else:
                # Formatação das células de dados
                if j in [0, 1, 5, 6]:  # Colunas com texto que precisa de alinhamento à esquerda
                    cell.alignment = cell_alignment
                else:  # Colunas com valores
                    cell.alignment = center_alignment

                # Destaca a linha recomendada
                if data[i][0].startswith("Profissional"):
                    cell.fill = highlight_fill

    # Deixa o texto em negrito nas células de endereço e custo total
    for row in range(start_row + 2, start_row + len(data)):
        ws.cell(row=row, column=2).font = bold_font
        ws.cell(row=row, column=5).font = bold_font

    # --- AJUSTE DE LARGURA DAS COLUNAS E ALTURA DAS LINHAS ---
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 45
    ws.column_dimensions['G'].width = 45

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    for i in range(start_row, start_row + len(data) + 1):
        ws.row_dimensions[i].height = 70  # Aumenta a altura para o texto caber

    # --- SALVAR O ARQUIVO ---
    try:
        wb.save(filename)
        print(f"Planilha '{filename}' gerada com sucesso!")
    except PermissionError:
        print(f"ERRO: Feche o arquivo '{filename}' se ele estiver aberto e tente novamente.")


# Executa a função para criar a planilha
if __name__ == '__main__':
    gerar_planilha_profissional()